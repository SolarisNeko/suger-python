import openpyxl


class ExcelUtils:
    @staticmethod
    def serialize(sheet, skip_rows=0) -> list:
        """
        将指定的 sheet 对象序列化为一个列表。
        skip_rows: 跳过的行数。
        """
        result = []
        for row in sheet.iter_rows(min_row=skip_rows + 1, values_only=True):
            result.append(list(row))
        return result

    @staticmethod
    def deserialize(sheet, dataList, skip_rows=0) -> None:
        """
        将一个 data 列表反序列化为指定的 sheet 对象。
        sheet: 目标 sheet 对象。
        data: 用于反序列化的 ${你的数据列表}
        skip_rows: 跳过的行数。
        """
        for i, row in enumerate(dataList):
            if i < skip_rows:
                continue
            sheet.append(row)

    @staticmethod
    def load_workbook(file_path):
        """
        加载指定的 Excel 文件。
        """
        return openpyxl.load_workbook(file_path)

    @staticmethod
    def save_workbook(workbook, file_path):
        """
        保存指定的 Excel 文件。
        """
        workbook.save(file_path)

    @staticmethod
    def get_sheet_by_name(workbook, sheet_name):
        """
        获取指定名称的 sheet 对象。
        """
        return workbook[sheet_name]

    @staticmethod
    def create_sheet(workbook, sheet_name):
        """
        创建一个新的 sheet 对象。
        """
        return workbook.create_sheet(title=sheet_name)
